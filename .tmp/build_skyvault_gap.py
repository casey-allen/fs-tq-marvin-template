#!/usr/bin/env python3
"""SkyVault Gap Analysis workbook — reconstructed per gap-analysis skill methodology
(bundled checklist/SOW/build script NOT installed; fallback mode, noted on Cover)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/casey/git/obsidian/Projects/SkyVault-VDR/SkyVault-Gap-Analysis.xlsx"
BLACK="000000"; WHITE="FFFFFF"; DARK="1A1A1A"; MED="595959"; LIGHT="F0F0F0"; ALT="F8F8F8"
def side(c="A0A0A0"): return Side(style="thin", color=c)
def thin(): return Border(left=side(),right=side(),top=side(),bottom=side())
def fill(c): return PatternFill("solid", fgColor=c)
def af(sz=9,b=False,color=BLACK,i=False): return Font(name="Arial",size=sz,bold=b,color=color,italic=i)
def wa(h="left"): return Alignment(horizontal=h,vertical="top",wrap_text=True)

PROJECT="Project SkyVault — Temple Strategic Power Offtake"
REVIEW="June 15, 2026"
VDR="Google Drive: BKV / Project SkyVault VDR (folder 1M4xm6...) — 9 top folders + FS Diligence"

PRIO={"Critical":0,"High":1,"Standard":2}
STAT={"Missing":0,"Stale":1,"Partial":2,"Confirmed":3,"N/A":4}

# item, discipline, status, source_file, notes, priority
gaps=[
 # Power - Interconnection / Transmission
 ("Generation Interconnection Agreement (existing units)","Power — Interconnection","Confirmed","6.1.4 T1_SGIA_Final","Executed Oncor<->Panda Temple Power SGIA 6/21/2012, GIR 10INR0020, Knob Creek 345 kV. T2 SGIA + amendment also present (scanned).","Critical"),
 ("Data-center LOAD interconnection agreement / Final FEA","Power — Interconnection","Missing","—","Only an interim, non-binding Oncor Substation/Transmission Facility Extension Agreement (1,500 MW, $6.5M placeholder). Oncor not obligated to build; Final FEA gated on completed studies.","Critical"),
 ("ERCOT/Oncor load study (LLIS results / SIS / Facilities Study)","Power — Interconnection","Missing","—","LLIS results not yet received. Only BKV's self-performed EPE screening study (3.9.1).","Critical"),
 ("Network upgrade scope & lead times","Power — Interconnection","Partial","3.9.1 EPE Transmission Analysis","EPE study lists ~13 limiters to resolve for 750 MW (345 kV Temple Sw-Knob Creek binds, SF 74-75%). Oncor scope/lead-times undefined.","High"),
 ("CIAC / transmission security cost schedule","Power — Interconnection","Missing","3.7.1 Interim FEA","$6.5M interim deposit is explicitly a placeholder; Final FEA upgrade cost TBD.","High"),
 ("Transmission deliverability / load-flow study","Power — Interconnection","Partial","3.9.2 Export Transmission Analysis","Self-performed EPE 2028-Summer snapshot; export capped 1,912 MW; EPE recommends an 8760-hr congestion study (not in room).","High"),
 ("Single-line diagrams (POI / switchyard)","Power — Interconnection","Partial","3.1.x / 3.8.x","Present but scanned image PDFs; not machine-readable. 1.5-breaker 345 kV per SGIA.","Standard"),
 # Power - Generation / BTM
 ("Generation operational / reliability data (GADS)","Power — Generation","Confirmed","3.3.1 / 3.3.2 GADS 2014-2025","T1/T2 GADS: EAR 89-93% (2025), EFORd <1%. Capacity factor only 54-60% (never baseload).","High"),
 ("RAM / availability study (supports 99.99% claim)","Power — Generation","Missing","—","Scheduled for modular; no RAM study in room. Units have not operated at 99.99% continuous output.","Standard"),
 ("Temple III generation IA / SGIA amendment (CCGT3)","Power — Generation","Missing","—","No T3 generation interconnection agreement found; export congestion caps deliverable ~426 MW absent upgrades.","High"),
 ("Generation equipment procurement (turbine RFA/PO)","Power — Generation","Confirmed","3.12.1 Schedule","GE Vernova 7HA.02 RFA executed 1/30/2026 (Payment 1 done); modular 34x INNIO Jenbacher ECO33 (~100 MW).","High"),
 ("Independent Engineer's report","Power — Generation","Missing","—","Scheduled in P6 for 2027; not yet produced. No third-party validation of availability/heat-rate/cost.","Standard"),
 ("O&M agreements (existing units)","Power — Generation","Confirmed","3.6.1 / 3.6.2","Temple I & II O&M (operator CAMS) execution versions; EDF EMA; budgets .xlsm unreadable.","Standard"),
 ("Siemens LTSA contract (existing units)","Power — Generation","Missing","—","Referenced throughout; the LTSA agreement itself is not in the room.","Standard"),
 ("Integrated project schedule","Power — Generation","Confirmed","3.12.1","P6-style schedule (~1,180 lines), dated 4/1/2026; Modular SC 10/2027, PUN ISD 5/2029, T3 SC 8/2030.","High"),
 ("Project cost estimate / capex","Power — Generation","Confirmed","3.13.1","Modular ~$0.9-1.06B, Temple III ~$1.18-1.35B, PUN ~$160-285M. No CIAC line; no who-pays-what allocation.","High"),
 # Power - Substation / PUN
 ("Interim substation facility extension agreement","Power — Substation","Confirmed","3.7.1 / 6.4.2","Executed Oncor Interim FEA, 1,500 MW, $6.5M security (interim/non-construction).","High"),
 ("Customer substation design (FEL)","Power — Substation","Partial","6.4.3 Load Questionnaire","8-transformer 345/34.5 kV customer substation; Moore switchyard at FEL-2. Not built.","Standard"),
 ("ERCOT PUN registration (Nodal Protocol 10.3.2.3)","Power — Substation","Missing","3.1.2 PUN Plan","Only a Dashiell feasibility proposal ($50k, Jun 2025); PUN not formally established/registered with ERCOT.","High"),
 ("Shared Facilities Agreement (substation co-ownership)","Power — Substation","Missing","—","Central to PUN co-ownership; referenced in all term sheets; not in the room.","Critical"),
 ("Breaker configuration / ratings","Power — Substation","Partial","3.8.1 / 3.8.2","Scanned images; 1.5-breaker 345 kV per SGIA; kA/A ratings not extractable.","Standard"),
 # Commercial
 ("Executed RPSA / Power Service Agreement","Commercial","Missing","2.3.1 Form PSA","Form only; all pricing/security fields blank. The commercial deal is uncommitted.","Critical"),
 ("Power Service Agreement term sheets","Commercial","Confirmed","2.1.1-2.1.3","Temple 1&2, Temple 3, Modular term sheets; non-binding; pricing points to Offtaker Questionnaire.","High"),
 ("PPA scenario / pricing analysis","Commercial","Confirmed","2.5.1","Jan 2026; 4 scenarios (Modular+PUN); unlevered, 100% CF; Temple III load = 0 in all.","Standard"),
 ("Offtaker Questionnaire (indicative pricing source)","Commercial","Missing","—","All term sheets point to it for $/MWh; not in the room.","High"),
 ("Land PSA (DC parcel form)","Commercial","Partial","2.4.1 Form PSA","Form land-sale agreement; economics blank; Seller repurchase right if buyer doesn't build.","High"),
 ("Buyer credit / security terms ($ amounts)","Commercial","Missing","2.x term sheets","Buyer must be investment-grade + 2-yr NPV performance security; ERCOT large-load security TBD; $ amounts blank; seller support 'to be discussed'.","High"),
 ("NDA","Commercial","Missing","—","Referenced in every confidentiality clause; not present.","Standard"),
 ("Carbon offering terms (CSG/REC/CCS)","Commercial","Partial","2.2.1-2.2.4","Marketing one-pagers; CCS study 2023 (Temple I only, stale); no SkyVault pricing or contractual carbon term.","Standard"),
 # Land & Site
 ("Land control — executed PSAs (DC parcels)","Land & Site","Partial","5.10.x PSAs","Most Bowie PSAs executed (Bratton/Medina/Fletcher/Barnhart/Shenkir/Lewis/Schwake); Area 3 (~389 ac) at LOI only.","Critical"),
 ("Plant land ownership documentation","Land & Site","Confirmed","5.5.2 Supplemental Ownership Report","Panda Temple Power / CXA Temple 2 / Panda Temple Power II / Oncor parcels; ~282-309 ac.","High"),
 ("Buildable / net-usable acreage quantification","Land & Site","Missing","5.8.12 Net Usable Area.kmz","Net usable locked in unreadable KMZ; not tabulated. Reduced by floodplain + >10% slope.","High"),
 ("Plant facilities lease (long-form)","Land & Site","Partial","5.5 Memorandum of Lease","Only a Memorandum of Lease recorded (Temple Generation SF, to 12/31/2058); long-form not provided.","Standard"),
 ("Expansion land (Area 3) control","Land & Site","Partial","5.4 Expanded Land Update","~389 ac at LOI; no PSAs yet.","Standard"),
 # Title
 ("Title commitment — DC parcels","Title","Partial","5.11.x","Per-parcel T-7 commitments issued; Lewis & Barnhart carry open Schedule C (spouse joinder, capacity, good-standing).","Critical"),
 ("Title commitment — plant","Title","Stale","5.5.1 Commitment for Title Insurance","Fidelity T-7 No. 280010230176, effective 5/29/2023 — >12 months; bring-down required.","High"),
 ("Lien / encumbrance review","Title","Partial","5.5.2","CLMG Corp $435M deed of trust on CXA Temple 2 land; numerous TP&L/Oncor easements; Gunter gas pipeline easement.","High"),
 ("Mineral rights / severance analysis","Title","Partial","5.7.3 Barnhart Mineral Opinion","Severed throughout (surface-only). Barnhart 1940 reservation opined expired 1960; surface waiver recommended.","High"),
 ("Grantor capacity resolution (Schwake life estate)","Title","Missing","5.10.1.7 / affidavits","Life-estate grantor age 90 with documented dementia; reliance on 2006 POA. No guardianship/judicial approval; title-insurability TBD.","High"),
 # Survey
 ("ALTA/NSPS survey (current, machine-readable)","Survey","Partial","5.2.x","Surveys present only as image PDFs/DWG; several PSAs note legal description 'to be updated upon survey'.","Critical"),
 ("Boundary / legal survey","Survey","Partial","5.2.x / 5.9.x","Some 2/2026 surveys (ACS / Lucko RPLS); tentative legal descriptions on Lewis/Bratton/Medina.","High"),
 ("Topographic survey / LIDAR","Survey","Partial","5.2.x / 5.8.6 Contour.kmz","LIDAR/slope DWGs + contour KMZ; not tabulated; >10% slope flagged on Bowie.","Standard"),
 # Environmental
 ("Phase I ESA — operating plant","Environmental","Stale","7.2.1 Terracon ESA","Dated 6/14/2023 — exceeds ASTM E1527-21 1-yr AAI window; refresh required for CERCLA defense.","Critical"),
 ("Phase I ESA — Temple III","Environmental","Confirmed","7.2.2 HDR ESA","HDR, 12/5/2025; 1 REC (arsenic GW), 2 HRECs (resolved). Current.","High"),
 ("Phase I ESA — Bowie / PUN parcels","Environmental","Partial","7.2.3-7.2.6","HDR done on Moore/Shenkir/BFL; PUN ROW has none; some Bowie parcels not yet done (ROE-dependent).","High"),
 ("Phase II ESA + PFAS groundwater sampling","Environmental","Missing","7.2.6 recommendation","HDR recommends Phase II (BFL: 1995 AFFF/PFAS, abandoned UST, lead; Moore: arsenic). Not performed; no sampling data, no NFA.","Critical"),
 ("Wetlands delineation (field)","Environmental","Missing","5.8.x desktop","Desktop only; Knob Creek (effluent-fed lower reach) likely jurisdictional WOTUS. Field delineation required.","High"),
 ("USACE 404 jurisdictional determination","Environmental","Missing","—","No JD; NWP 39/57 anticipated; IP (12+ mo) if Knob Creek rerouted.","High"),
 ("T&E species survey","Environmental","Partial","5.8.2 / HDR","IPaC desktop; no critical habitat; tricolored bat -> winter tree-clearing. No field survey.","Standard"),
 ("Cultural / archaeological (THC) survey","Environmental","Partial","5.8.x desktop","4 recorded sites overlap Temple III; 41BL1257/1258 unevaluated. §106 only with federal nexus.","High"),
 ("Abandoned UST closure (Parcel 397331)","Environmental","Missing","7.2.6","Unregistered 1930s diesel UST, never tightness-tested or closed; still in ground. GPR + Phase II needed.","High"),
 # Civil / Geotech / Flood
 ("Geotechnical report (DC footprint)","Civil / Geotech","Missing","—","No geotechnical study in VDR; compounded by >10% slope and terracing on much of the Bowie land.","Critical"),
 ("Floodplain determination / H&H analysis","Civil / Geotech","Missing","5.8.x HDR desktop","Knob Creek floodway+100yr+500yr bisects Temple III & SkyVault; HDR desktop only; no FEMA determination or CLOMR.","High"),
 ("Stormwater / drainage plan","Civil / Geotech","Missing","—","Not in VDR; facility planned zero-liquid-discharge.","Standard"),
 ("Site logistics / heavy-haul / laydown plan","Civil / Geotech","Missing","—","BNSF/Oncor/TxDOT 90-degree crossings noted; no consolidated route/laydown plan.","Standard"),
 # Zoning & Permitting
 ("Zoning confirmation / PZR","Zoning & Permitting","Partial","5.12.1 / 5.12.2 PZR","Temple III LI(PD) conforming; Bowie unincorporated ETJ (by-right vacant). Bell County issues no permits/CO.","Critical"),
 ("Annexation + PD-LI rezoning","Zoning & Permitting","Missing","5.1.1 / 5.1.2","Owner-initiated via City Council hearing; not filed. Required to unlock City of Temple permits.","High"),
 ("Permitting matrix","Zoning & Permitting","Partial","9.1 / 9.2 / 5.8.x","Four HDR preliminary desktop matrices; nothing submitted or obtained.","High"),
 ("Building / grading permits","Zoning & Permitting","Missing","—","Not submitted; gated on annexation; 'applicability to be confirmed by city'.","Standard"),
 ("Floodplain development permit (Bell County)","Zoning & Permitting","Missing","—","Required (Knob Creek); not obtained.","High"),
 ("Tax abatement agreement","Zoning & Permitting","Missing","—","Contemplated via annexation; not executed/quantified.","Standard"),
 # Air Quality
 ("Existing-unit air permit (NSR / PSD)","Air Quality","Confirmed","7.1.4 / 7.1.3","TCEQ NSR Permit 83503 + PSD PSDTX1111; 2024 actuals within MAERT; CEMS for NOx/CO.","High"),
 ("Title V (Federal Operating Permit)","Air Quality","Missing","—","Only NSR/PSD in room; 7.1.2 Air Quality Report is an unreadable scan.","Standard"),
 ("Modular standard air permit #183943","Air Quality","Missing","Q&A reference","Reported issued 5/22/2026 for 400 MW modular; document not in the room.","High"),
 ("Temple III PSD air permit application","Air Quality","Missing","—","Not applied; PSD applicability 'needs confirmation' (essentially certain for 750 MW CCGT).","High"),
 ("Emissions inventory (existing units)","Air Quality","Confirmed","7.1.1","2024 EI; 0 reportable emission events; NOx 202.9 / CO 296.4 TPY.","Standard"),
 # Water / Wastewater
 ("Water will-serve letter + committed volumes","Water / Wastewater","Missing","4.1.1","City of Temple reclaimed-water rate sheet only; no will-serve, no GPM/AF-yr.","High"),
 ("Water supply agreements (T1/T2)","Water / Wastewater","Partial","4.1.2 / 4.1.3","Scanned/unreadable; contracted plant volumes not confirmable.","Standard"),
 ("Cooling-water adequacy (Temple III + DC)","Water / Wastewater","Missing","—","No volumes; adequacy for added Temple III + DC cooling unassessed.","High"),
 ("Wastewater / ZLD plan","Water / Wastewater","Partial","matrices","ZLD facility (TPDES N/A); Bell County septic for unincorporated; not detailed.","Standard"),
 # Connectivity
 ("Fiber connectivity assessment","Connectivity","Partial","3.4.1 Altman Solon (excerpts)","5 diverse paths; FiberLight/Lumen/Uniti <1.5 mi. 'Limited Excerpts' only; full report + measured latency missing.","Standard"),
 ("Carrier agreements / IRU terms","Connectivity","Missing","—","Verbal carrier interest noted; no draft IRU/dark-fiber lease.","Standard"),
 # Insurance
 ("Insurance — existing plant (property/BI)","Insurance","Confirmed","3.5.1 / 3.5.2","Blanket property $685M; BI ~$242M; large subscription program (Aon).","Standard"),
 ("Builders-risk / DSU insurance (newbuild)","Insurance","Missing","—","Not in room; expected for a financed newbuild + modular.","Standard"),
 # ERCOT Regional Addendum (reconstructed)
 ("TFEA/DFEA (Transmission Facility Extension Agreement)","ERCOT Addendum","Partial","3.7.1 / 6.4.2","Interim FEA executed (1,500 MW, $6.5M); Final TFEA pending completed studies.","High"),
 ("TFEA/DFEA assignment instrument (distinct from agreement)","ERCOT Addendum","Missing","—","No separate assignment instrument in the room.","Standard"),
 ("REP / QSE / LSE registration (ERCOT market participation)","ERCOT Addendum","Confirmed","6.3.x","BKV-BPP Retail REP Cert 10323 (PUCT Docket 53943); QSE + LSE registrations + Std Form MP Agreement.","Standard"),
 ("ERCOT generation registration (existing units)","ERCOT Addendum","Confirmed","6.2.1 / 6.2.2","Temple I/II registered generation resources (scanned); units operating.","Standard"),
 ("SB6 large-load compliance package (ride-through / UFLS / curtailment plan)","ERCOT Addendum","Missing","3.7.1 Art. V","Texas SB6 large-load rules under development; explicit CP contemplating curtailment; no compliance package.","High"),
 ("Site Control Attestation / Duplicative-Load Disclosure (Oncor large-load)","ERCOT Addendum","Confirmed","6.4.1","Signed 11/7/2025 (Oncor >=75 MW load requirement).","Standard"),
 ("GCD groundwater permit (Clearwater UWCD)","ERCOT Addendum","Missing","—","Clearwater Underground Water Conservation District governs unincorporated groundwater; no permit/assessment.","Standard"),
 ("TX Large Data Center Sales/Use Tax exemption","ERCOT Addendum","Missing","—","Not applied or secured; material to project economics.","Standard"),
 # Justified N/A items
 ("Coastal / Zone VE flood analysis","Civil / Geotech","N/A","—","N/A — inland Central Texas (Bell County); no coastal/VE flood zone within study area. (Riverine Knob Creek floodplain covered separately.)","Standard"),
 ("Seismic risk assessment","Civil / Geotech","N/A","—","N/A — Bell County, TX is a low-seismic zone (USGS low PGA); no SDC concern for the BoD.","Standard"),
 ("ISO interconnection queue position confirmation","Power — Interconnection","N/A","—","N/A — ERCOT is non-FERC; existing generation is operational and the DC load is processed via Oncor FEA/LLIS, not a FERC interconnection queue.","Standard"),
]

# sort: discipline, priority, status(Missing first)
gaps_sorted = sorted(gaps, key=lambda g:(g[1], PRIO.get(g[5],9), STAT.get(g[2],9)))

files_reviewed=[
 ("1.1 Teaser / 1.2 CIM / 1.3 Process Letter","Commercial/Process","2025-11-19"),
 ("2.1.1-2.1.3 Term Sheets (T1&2, T3, Modular)","Commercial","2026"),
 ("2.2.1-2.2.4 Carbon (Capture Study, Credits, CSG)","Commercial","2023-2025"),
 ("2.3.1 Form Power Service Agreement","Commercial","2026"),
 ("2.4.1 Form PSA (DC land sale)","Commercial","2026"),
 ("2.5.1 PPA Scenario Analysis","Commercial","2026-01-30"),
 ("3.1.x Temple 1&2 PUN drawings / Dashiell PUN plan","Power","2025-2026"),
 ("3.2.x Temple III EPC Plan / Permitting Matrix / Schedule","Power","2026"),
 ("3.3.1/3.3.2 GADS Operational Data; 3.3.3 Eng. Deep Dive","Power","2014-2026"),
 ("3.5.1/3.5.2 Insurance (property, liability)","Power","2024"),
 ("3.6.1/3.6.2 O&M Agreements (CAMS)","Power","exec"),
 ("3.7.1 Interim Substation Facility Extension Agreement","Power","2025-11"),
 ("3.8.x Existing Plant Breaker Configs (scanned)","Power","—"),
 ("3.9.1/3.9.2 EPE Transmission Analysis + tables","Power","2026"),
 ("3.11.x Modular Generation design","Power","2026"),
 ("3.12.1 Skyvault P6 Schedule","Power","2026-04-01"),
 ("3.13.1 Temple Project Cost Estimate","Power","2026-04-10"),
 ("4.1.1 City of Temple Water Agreement; 4.1.2/4.1.3 (scanned)","Water","—"),
 ("4.2.1 Temple Gas Agreements","Power/Gas","—"),
 ("3.4.1 Altman Solon Fiber Assessment (excerpts)","Connectivity","2025-10"),
 ("5.2.x Site images / ALTA / LIDAR (image/CAD)","Survey","2026"),
 ("5.3/5.4 Land Updates","Land","2025-11/12"),
 ("5.5.x Title commitment, ownership report, deeds","Title","2023"),
 ("5.6.1 Project Area Outlines (KMZ)","Survey","2026-01"),
 ("5.7.3 Barnhart Mineral Letter Opinion","Title","2026-03-11"),
 ("5.8.x HDR Constraints (T3/Moore/Bowie/PUN) + KMZ","Civil/Env","2026-02"),
 ("5.9.x Title Report (LTC spreadsheets)","Title","2025-2026"),
 ("5.10.x PSAs (Bratton/Medina/Fletcher/Barnhart/Shenkir/Lewis/Schwake)","Land","2025-2026"),
 ("5.11.x Title Commitments (per parcel)","Title","2026-01/02"),
 ("5.12.1/5.12.2 PZR Reports (Komar, Bowie Block)","Zoning","2026-03-13"),
 ("6.1.x Temple 1&2 SGIA + amendment","Power","2012"),
 ("6.2.x ERCOT Generation Registration","Power","—"),
 ("6.3.x REP/QSE/LSE registration (BKV-BPP Retail)","Power","2022"),
 ("6.4.x Site Control Attestation, Interim FEA, 1500 MW Load Questionnaire","Power","2025-11"),
 ("7.1.1-7.1.4 Emissions Inventory, MAERT, Permit Special Conditions","Environmental","2024-2025"),
 ("7.2.1-7.2.6 Phase I ESAs (Terracon plant; HDR T3/Moore/Shenkir/BFL)","Environmental","2023-2026"),
 ("9.1 Moore Permitting Matrix; 9.2 Temple Permitting Matrix","Zoning","2026-04"),
 ("FS prior diligence: Questions log, BKV Q&A, Gap Analysis","Commercial","2026-05/06"),
]

from collections import Counter
cnt = Counter(g[2] for g in gaps_sorted)

wb=Workbook()
# Sheet 1: Cover
ws=wb.active; ws.title="Cover"
ws.merge_cells("A1:D1"); c=ws["A1"]; c.value=f"DILIGENCE GAP ANALYSIS — {PROJECT}"
c.font=Font(name="Arial",size=13,bold=True,color=WHITE); c.fill=fill(DARK); c.alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[1].height=26
rows=[("Review Date",REVIEW),("VDR Source",VDR),
 ("Summary Counts",f"{cnt.get('Confirmed',0)} Confirmed  |  {cnt.get('Partial',0)} Partial  |  {cnt.get('Missing',0)} Missing  |  {cnt.get('N/A',0)} N/A  |  {cnt.get('Stale',0)} Stale   (Total {len(gaps_sorted)})"),
 ("Files Reviewed",f"{len(files_reviewed)} document sets across 9 VDR folders + FS prior diligence")]
r=3
for k,v in rows:
    ws.cell(row=r,column=1,value=k).font=af(10,b=True); ws.merge_cells(f"B{r}:D{r}")
    ws.cell(row=r,column=2,value=v).font=af(10); ws.cell(row=r,column=2).alignment=wa()
    ws.row_dimensions[r].height=30 if k=="Summary Counts" else 18; r+=1
r+=1
ws.cell(row=r,column=1,value="Methodology").font=af(10,b=True); r+=1
method=("This is a diligence completeness check against Fluidstack's pre-design SOW and site-selection requirements. "
 "Each requirement is classed Confirmed (present, complete, current), Partial (present but incomplete/scope-limited/desktop), "
 "Missing (not in VDR — the default), N/A (justified inapplicable), or Stale (>12 months or pre-dates a material change). "
 "It scores nothing and rates no risk — see the companion vdr-review workbook (SkyVault-VDR-Tables.xlsx) for the scored review, "
 "82-item risk matrix, and recommendation.")
ws.merge_cells(f"A{r}:D{r}"); ws.cell(row=r,column=1,value=method).font=af(9,i=True); ws.cell(row=r,column=1).alignment=wa(); ws.row_dimensions[r].height=70; r+=2
note=("FALLBACK MODE: The gap-analysis skill's bundled checklist (diligence_checklist.md), ERCOT regional addendum, static SOW "
 "stage-gate file (sow_stage_gates.json), and build_excel.py were not installed in this environment. This matrix was reconstructed "
 "from the skill methodology + ERCOT domain knowledge + the shared VDR read. The static 'SOW Reference' tab is therefore omitted; "
 "SOW Ref numbers are not mapped. ERCOT addendum items were reconstructed from memory and should be reconciled against the canonical "
 "addendum when installed.")
ws.merge_cells(f"A{r}:D{r}"); cc=ws.cell(row=r,column=1,value=note); cc.font=Font(name="Arial",size=9,bold=True,color="000000"); cc.fill=fill("F0F0F0"); cc.alignment=wa(); ws.row_dimensions[r].height=80
for col,w in zip("ABCD",[20,40,30,30]): ws.column_dimensions[col].width=w

# Sheet 2: Files Reviewed
ws2=wb.create_sheet("Files Reviewed")
ws2.merge_cells("A1:C1"); c=ws2["A1"]; c.value=f"FILES REVIEWED — {PROJECT}"; c.font=Font(name="Arial",size=12,bold=True,color=WHITE); c.fill=fill(DARK); c.alignment=Alignment(horizontal="left",vertical="center"); ws2.row_dimensions[1].height=22
hdr=["File / Document Set","Discipline","Date"]
for i,h in enumerate(hdr,1):
    cell=ws2.cell(row=2,column=i,value=h); cell.font=af(9,b=True,color=WHITE); cell.fill=fill(MED); cell.border=thin(); cell.alignment=Alignment(horizontal="center")
for ri,(n,d,dt) in enumerate(files_reviewed,3):
    bg=ALT if ri%2 else WHITE
    for ci,val in enumerate((n,d,dt),1):
        cell=ws2.cell(row=ri,column=ci,value=val); cell.font=af(9); cell.fill=fill(bg); cell.border=thin(); cell.alignment=wa()
    ws2.row_dimensions[ri].height=16
for col,w in zip("ABC",[62,22,14]): ws2.column_dimensions[col].width=w

# Sheet 3: Gap Matrix
ws3=wb.create_sheet("Gap Matrix")
ws3.merge_cells("A1:F1"); c=ws3["A1"]; c.value=f"GAP MATRIX — {PROJECT}"; c.font=Font(name="Arial",size=12,bold=True,color=WHITE); c.fill=fill(DARK); c.alignment=Alignment(horizontal="left",vertical="center"); ws3.row_dimensions[1].height=22
hdr=["Item","Discipline","Status","Source File","Notes","Priority"]
for i,h in enumerate(hdr,1):
    cell=ws3.cell(row=2,column=i,value=h); cell.font=af(9,b=True,color=WHITE); cell.fill=fill(MED); cell.border=thin(); cell.alignment=Alignment(horizontal="center",vertical="center");
ws3.row_dimensions[2].height=18
for ri,(item,disc,stat,src,notes,prio) in enumerate(gaps_sorted,3):
    bg=ALT if ri%2 else WHITE
    vals=(item,disc,stat,src,notes,prio)
    for ci,val in enumerate(vals,1):
        cell=ws3.cell(row=ri,column=ci,value=val); cell.fill=fill(bg); cell.border=thin(); cell.alignment=wa()
        cell.font=af(9,b=(ci in (3,6)))
    ws3.row_dimensions[ri].height=34
ws3.freeze_panes="A3"
for col,w in zip("ABCDEF",[40,20,11,26,62,11]): ws3.column_dimensions[col].width=w

wb.save(OUT)
print("Saved:",OUT)
print("Counts:",dict(cnt),"Total",len(gaps_sorted))
